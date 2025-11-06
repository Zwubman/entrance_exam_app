import io
import os
import uuid
import textwrap
from PIL import Image
from app.config.setting import settings
from openpyxl import load_workbook

def extract_xls_data(file_bytes: bytes):
    file_stream = io.BytesIO(file_bytes)
    extracted_payloads = []

    try:
        workbook = load_workbook(filename=file_stream, data_only=True)
        file_type = "xlsx"
    except Exception:
        try:
            import xlrd
            file_stream.seek(0)
            workbook = xlrd.open_workbook(file_contents=file_stream.read())
            file_type = "xls"
        except ImportError:
            print("xlrd not installed for .xls file support")
            return []
        except Exception as e:
            print(f"Unsupported Excel format: {e}")
            return []

    for sheet_name in workbook.sheetnames if file_type == "xlsx" else workbook.sheet_names():
        page_texts, page_images = [], []
        
        try:
            sheet = workbook[sheet_name] if file_type == "xlsx" else workbook.sheet_by_name(sheet_name)

            if file_type == "xlsx":
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell).strip() for cell in row if cell is not None and str(cell).strip()])
                    if row_text:
                        sentences = textwrap.wrap(row_text, width=settings.CHUNK_LENGTH)
                        page_texts.extend(sentences)
                        
                # Extract images from xlsx
                for image in sheet._images:
                    try:
                        img_data = image._data()
                        img = Image.open(io.BytesIO(img_data))
                        image_name = f"{sheet_name}_{uuid.uuid4().hex}.png"
                        image_path = os.path.join(settings.UPLOADS_DIR, image_name)
                        img.save(image_path, "PNG")
                        page_images.append(image_path)
                    except Exception as e:
                        print(f"Failed to extract image from {sheet_name}: {e}")
                        
            else:  # xls format
                for row_idx in range(sheet.nrows):
                    row_values = [str(v).strip() for v in sheet.row_values(row_idx) if v and str(v).strip()]
                    if row_values:
                        row_text = " ".join(row_values)
                        sentences = textwrap.wrap(row_text, width=settings.CHUNK_LENGTH)
                        page_texts.extend(sentences)
                        
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")
            continue

        extracted_payloads.append({
            "texts": page_texts,
            "images": page_images
        })

    return extracted_payloads